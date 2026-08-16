import os
from typing import List, Dict, Any
from openpyxl import load_workbook
from utils.logger import get_logger

logger = get_logger(__name__)

def load_excel_data(file_path: str, sheet_name: str) -> List[Dict[str, Any]]:
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        return []

    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        logger.warning(f"Sheet {sheet_name} not found.")
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return []

    header = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    data_rows: List[Dict[str, Any]] = []

    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
            
        row_dict = {header[i]: row[i] for i in range(min(len(header), len(row)))}
        
        # Only run cases where the Enabled column is 'Y'
        enabled_val = str(row_dict.get("Enabled", "")).strip().upper()
        if enabled_val != "Y":
            continue

        data_rows.append(row_dict)
        
    return data_rows